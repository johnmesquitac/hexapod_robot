import cv2
import math
from digi.xbee.devices import XBeeDevice
import matplotlib.pyplot as plt
import time
from openpyxl import Workbook

def AdicionarPlanilha():
    global ws
    global wb
    global robot_center_x 
    global robot_center_y 
    global robot_inclination 
    global target_x 
    global target_y 
    global target_inclination
    global split_time
    global cont_linhas_planilha    
    linha = str(cont_linhas_planilha)
    ws['A' + linha] = robot_center_x
    ws['B' + linha] = robot_center_y
    ws['C' + linha] = robot_inclination
    ws['D' + linha] = target_x
    ws['E' + linha] = target_y
    ws['F' + linha] = target_inclination
    ws['G' + linha] = split_time
    cont_linhas_planilha += 1

def CriarPlanilha():
    global ws
    global wb
    wb = Workbook()
    ws = wb.active
    ws.title = 'Execucao Robo'
    ws = wb['Execucao Robo']
    ws['A1'] = 'Robo X'
    ws['B1'] = 'Robo Y'
    ws['C1'] = 'Inclinacao Robo'
    ws['D1'] = 'Waypoint X'
    ws['E1'] = 'Waypoint Y'
    ws['F1'] = 'Inclinacao Waypoint'
    ws['G1'] = 'Tempo'

def VerificaPosicao():
    global x_red 
    global y_red 
    global x_blue  
    global y_blue 
    global robot_center_x 
    global robot_center_y 
    global robot_inclination 
    global target_x 
    global target_y 
    global target_inclination
    global List_X
    global List_Y
    global start_time
    global split_time
    ret_val, img = cam.read()
    img_filter = cv2.GaussianBlur(img.copy(), (3, 3), 0)
    img_filter = cv2.cvtColor(img_filter, cv2.COLOR_BGR2HSV)
    img_binary_red = cv2.inRange(img_filter.copy(), THRESHOLD_LOW_RED, 
	THRESHOLD_HIGH_RED)
    img_binary_blue = cv2.inRange(img_filter.copy(), THRESHOLD_LOW_BLUE, 
	THRESHOLD_HIGH_BLUE)
    img_binary_red = cv2.dilate(img_binary_red, None, iterations = 1)
    img_binary_blue = cv2.dilate(img_binary_blue, None, iterations = 1)

	#Encontrar círculo vermelho 
    img_contours = img_binary_red.copy()
    contours = cv2.findContours(img_contours, cv2.RETR_EXTERNAL, \
        cv2.CHAIN_APPROX_SIMPLE)[-2]
    center = None
    radius = 0
    if len(contours) > 0:
        c = max(contours, key=cv2.contourArea)
        ((x_red, y_red), radius) = cv2.minEnclosingCircle(c)
        M = cv2.moments(c)
        if M["m00"] > 0:
            center = (int(M["m10"] / M["m00"]), 
			int(M["m01"] / M["m00"]))
            if radius < MIN_RADIUS:
                center = None
    if center != None:
        cv2.circle(img, center, int(round(radius)), (0, 255, 0))

	#Encontrar círculo azul
    img_contours = img_binary_blue.copy()
    contours = cv2.findContours(img_contours, cv2.RETR_EXTERNAL, \
        cv2.CHAIN_APPROX_SIMPLE)[-2]
    center = None
    radius = 0
    if len(contours) > 0:
        c = max(contours, key=cv2.contourArea)
        ((x_blue, y_blue), radius) = cv2.minEnclosingCircle(c)
        M = cv2.moments(c)
        if M["m00"] > 0:
            center = (int(M["m10"] / M["m00"]), 
			int(M["m01"] / M["m00"]))
            if radius < MIN_RADIUS:
                center = None
    if center != None:
        cv2.circle(img, center, int(round(radius)), (0, 255, 0))

	#Exibindo resultados    
    x_red =  x_red * dist_x / CAMERA_WIDTH  
    y_red =  dist_y - (y_red * dist_y / CAMERA_HEIGHT) 
    x_blue = x_blue * dist_x / CAMERA_WIDTH 
    y_blue = dist_y - (y_blue * dist_y / CAMERA_HEIGHT)     
    robot_center_x = int( (x_blue + x_red) / 2 )
    robot_center_y = int( (y_blue + y_red) / 2 )
    List_X.append(robot_center_x)
    List_Y.append(robot_center_y)    
    robot_inclination = 
	round ( math.atan2(x_blue - x_red, y_blue - y_red), 2 )
    target_inclination = 
	round ( math.atan2(target_x - robot_center_x, 
	target_y - robot_center_y) , 2)     
	
	#Atualizando imagem da camera
    cv2.imshow('webcam', img) 
    cv2.waitKey(3)
    message = str(robot_center_x) + "," + str(robot_center_y) 
	+ ";" + str(robot_inclination) + ";" + str(target_x) 
	+ "," + str(target_y) + ";" + str(target_inclination)
    split_time = time.time() - start_time
    AdicionarPlanilha()
    print ("Enviado para o Robo: " + message)
    return message

def main():
    global target_x
    global target_y
    global List_Target_X
    global List_Target_Y
    global List_X
    global List_Y
    global start_time
    global split_time
    global wb
    global cont_linhas_planilha
    global ws    
    device = XBeeDevice(PORT, BUS)
    device.open()
    i = 0
    target = targets[i]
    target_x = int(target.split(',')[0])
    target_y = int(target.split(',')[1])
    List_Target_X.append(target_x)
    List_Target_Y.append(target_y)
    start_time = time.time()
    while True:
        print('Alvo: ' + str(target))
        if target is not None:  
            robot_info = device.read_data()            
            while robot_info is None:
                robot_info = device.read_data()
            robot_info = robot_info.data.decode()
            print ('Informacao Recebida do Robo: ' + str(robot_info))
            if 'Posicao' in robot_info:
                device.send_data_broadcast(VerificaPosicao())
            elif 'Cheguei' in robot_info:
                ws['H' + str(cont_linhas_planilha - 1)] = 'Cheguei'
                if i < len(targets) - 1:
                    i += 1
                    target = targets[i]                    
                    target_x = int(target.split(',')[0])
                    target_y = int(target.split(',')[1])
                    List_Target_X.append(target_x)
                    List_Target_Y.append(target_y)
                    device.send_data_broadcast(VerificaPosicao())
                else:
                    break                
            wb.save('relatório.xlsx')
            print ("\n\n")
    device.close()
    print ("\n\n")
    print ("O robo chegou a todos os alvos. Aplicacao encerrada.")
    plt.plot(List_X, List_Y, 'r--', 
	List_Target_X, List_Target_Y, 'bo', 
	List_X[0], List_Y[0], 'go')
    plt.axis([0, dist_x, 0, dist_y])
    plt.show()

#Parametros
# Azul
THRESHOLD_LOW_BLUE = (80, 140, 87);
THRESHOLD_HIGH_BLUE = (120, 240, 130); 
# Vermelho
THRESHOLD_LOW_RED = (0, 130, 120);
THRESHOLD_HIGH_RED = (15, 240, 180);
# Distancia visíveis nos eixos
dist_x = 277 
dist_y = 209
# Resolucao desejada
CAMERA_WIDTH = 1024
CAMERA_HEIGHT = 768
# Raio mínimo para o círculo de contorno
MIN_RADIUS = 2
PORT = "COM25"
BUS = 9600
# Inicializacao de variaveis globais
x_red = 0 
y_red = 0
x_blue = 0 
y_blue = 0
robot_center_x = 0
robot_center_y = 0
robot_inclination = 0
target_x = 0
target_y = 0
target_inclination = 0
start_time = 0
split_time = 0
cont_linhas_planilha = 2
ws = None
wb = None
List_X = []
List_Y = []
List_Target_X = []
List_Target_Y = []
#Lista de Waypoints
targets = ['210,70', '200,140', '150,170','100,170','50,50']
# Inicializacao da camera
cam = cv2.VideoCapture(1)
# Define a resolucao escolhida para a imagem
cam.set(cv2.CAP_PROP_FRAME_WIDTH, CAMERA_WIDTH)
cam.set(cv2.CAP_PROP_FRAME_HEIGHT, CAMERA_HEIGHT)
print ("Camera Inicializada")
print ("Monitoramento do Robo Ativo")
CriarPlanilha()
main()
